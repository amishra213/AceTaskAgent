"""
Excel Sub-Agent for handling Excel file operations.

Capabilities:
- Read Excel files and extract data
- Create Excel workbooks with formatted sheets
- Write data to specific cells/ranges
- Perform data aggregation and formatting
- Create charts and pivot tables

Migration Status: Week 7 Day 1 - Dual Format Support
- Supports both legacy dict and standardized AgentExecutionRequest/Response
- Maintains 100% backward compatibility
- Publishes SystemEvent on completion for event-driven workflows
"""

from typing import Optional, List, Dict, Any, Union
from pathlib import Path
from datetime import datetime
import json
import time

from task_manager.utils.logger import get_logger

# Import standardized schemas and utilities (Week 1-2 implementation)
from task_manager.models import (
    AgentExecutionRequest,
    AgentExecutionResponse,
    create_system_event,
    create_error_response
)
from task_manager.utils import (
    auto_convert_response,
    validate_agent_execution_response,
    exception_to_error_response,
    InvalidParameterError,
    ExcelOperationError,
    MissingDependencyError,
    wrap_exception
)
from task_manager.core.event_bus import get_event_bus

logger = get_logger(__name__)


class ExcelAgent:
    """
    Sub-agent for Excel file operations.
    
    This agent handles all Excel-related tasks:
    - Reading and extracting data from Excel files
    - Creating new Excel workbooks
    - Writing data to sheets
    - Data aggregation and formatting
    - Chart creation
    """
    
    def __init__(self):
        """Initialize Excel Agent with dual-format support."""
        self.agent_name = "excel_agent"
        self.supported_operations = [
            "read",
            "create",
            "write",
            "append",
            "format",
            "aggregate",
            "create_chart"
        ]
        
        # Initialize event bus for event-driven workflows
        self.event_bus = get_event_bus()
        
        logger.info("Excel Agent initialized with dual-format support")
        self._check_dependencies()
    
    
    def _check_dependencies(self):
        """Check if required Excel libraries are installed."""
        try:
            import openpyxl
            self.excel_lib = "openpyxl"
            logger.debug("Using openpyxl for Excel operations")
        except ImportError:
            try:
                import pandas as pd
                self.excel_lib = "pandas"
                logger.debug("Using pandas for Excel operations")
            except ImportError:
                logger.warning(
                    "No Excel library found. Install with: "
                    "pip install openpyxl (or pip install pandas)"
                )
                self.excel_lib = None
    
    
    def read_excel(
        self,
        file_path: str,
        sheet_name: Union[str, int, None] = None,
        header: int = 0,
        index_col: Optional[int] = None
    ) -> Dict[str, Any]:
        """
        Read and extract data from an Excel file.
        
        Args:
            file_path: Path to the Excel file
            sheet_name: Name or index of sheet to read (None = all sheets)
            header: Row number to use as column names
            index_col: Column number to use as index
        
        Returns:
            Dictionary with extracted data
        """
        try:
            if not self.excel_lib:
                return {
                    "success": False,
                    "error": "Excel library not installed",
                    "file": file_path
                }
            
            file_path_obj = Path(file_path)
            if not file_path_obj.exists():
                return {
                    "success": False,
                    "error": f"File not found: {file_path}",
                    "file": str(file_path_obj)
                }
            
            logger.info(f"Reading Excel file: {file_path}")
            
            if self.excel_lib == "pandas":
                import pandas as pd  # noqa: F401
                
                try:
                    # Read Excel file
                    if sheet_name is None:
                        # Read all sheets
                        excel_file = pd.read_excel(
                            str(file_path_obj),
                            sheet_name=None,
                            header=header,
                            index_col=index_col
                        )
                        
                        sheets_data = {}
                        for sheet_name_key, df in excel_file.items():
                            sheets_data[sheet_name_key] = {
                                "rows": len(df),
                                "columns": list(df.columns),
                                "data": df.to_dict('records'),
                                "summary": {
                                    "shape": df.shape,
                                    "dtypes": df.dtypes.to_dict(),
                                    "missing_values": df.isnull().sum().to_dict()
                                }
                            }
                        
                        logger.info(f"Read {len(sheets_data)} sheets from Excel file")
                        
                        return {
                            "success": True,
                            "file": str(file_path),
                            "sheets": list(sheets_data.keys()),
                            "data": sheets_data,
                            "total_sheets": len(sheets_data),
                            "read_at": datetime.now().isoformat()
                        }
                    else:
                        # Read specific sheet
                        df = pd.read_excel(
                            str(file_path_obj),
                            sheet_name=sheet_name,
                            header=header,
                            index_col=index_col
                        )
                        
                        logger.info(f"Read sheet '{sheet_name}' with {len(df)} rows")
                        
                        return {
                            "success": True,
                            "file": str(file_path),
                            "sheet": str(sheet_name),
                            "rows": len(df),
                            "columns": list(df.columns),
                            "data": df.to_dict('records'),
                            "summary": {
                                "shape": df.shape,
                                "dtypes": df.dtypes.to_dict(),
                                "missing_values": df.isnull().sum().to_dict()
                            },
                            "read_at": datetime.now().isoformat()
                        }
                
                except Exception as e:
                    logger.error(f"Error reading with pandas: {str(e)}")
                    raise
            
            elif self.excel_lib == "openpyxl":
                from openpyxl import load_workbook  # noqa: F401
                
                workbook = load_workbook(str(file_path_obj))
                
                if sheet_name is None:
                    # Read all sheets
                    sheets_data = {}
                    
                    for sheet_name_key in workbook.sheetnames:
                        ws = workbook[sheet_name_key]
                        data = []
                        
                        for row in ws.iter_rows(values_only=True):
                            data.append(row)
                        
                        sheets_data[sheet_name_key] = {
                            "rows": len(data),
                            "columns": data[0] if data else [],
                            "data": [dict(zip(data[0], row)) for row in data[1:]] if len(data) > 1 else [],
                            "raw_data": data
                        }
                    
                    logger.info(f"Read {len(sheets_data)} sheets from Excel file")
                    
                    return {
                        "success": True,
                        "file": str(file_path),
                        "sheets": list(sheets_data.keys()),
                        "data": sheets_data,
                        "total_sheets": len(sheets_data),
                        "read_at": datetime.now().isoformat()
                    }
                else:
                    # Read specific sheet
                    if isinstance(sheet_name, int):
                        ws = workbook.worksheets[sheet_name]
                    else:
                        ws = workbook[sheet_name]
                    
                    data = []
                    for row in ws.iter_rows(values_only=True):
                        data.append(row)
                    
                    logger.info(f"Read sheet with {len(data)} rows")
                    
                    return {
                        "success": True,
                        "file": str(file_path),
                        "sheet": str(sheet_name),
                        "rows": len(data),
                        "columns": data[0] if data else [],
                        "data": [dict(zip(data[0], row)) for row in data[1:]] if len(data) > 1 else [],
                        "raw_data": data,
                        "read_at": datetime.now().isoformat()
                    }
            
            return {
                "success": False,
                "error": "No Excel library available",
                "file": str(file_path)
            }
        
        except Exception as e:
            logger.error(f"Error reading Excel file: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "file": str(file_path)
            }
    
    
    def create_excel(
        self,
        output_path: str,
        sheets: Dict[str, List[List[Any]]],
        title: Optional[str] = None,
        auto_format: bool = True
    ) -> Dict[str, Any]:
        """
        Create an Excel workbook with multiple sheets.
        
        Args:
            output_path: Path where Excel file will be saved
            sheets: Dictionary of sheet_name -> data (list of rows)
            title: Workbook title
            auto_format: Whether to auto-format the sheets
        
        Returns:
            Dictionary with creation result
        """
        try:
            logger.info("=" * 80)
            logger.info("[EXCEL CREATE_EXCEL] Starting Excel file creation")
            logger.info("=" * 80)
            logger.info(f"[CREATE] Output path: '{output_path}'")
            logger.info(f"[CREATE] Output path type: {type(output_path)}")
            logger.info(f"[CREATE] Output path length: {len(output_path) if output_path else 0}")
            logger.info(f"[CREATE] Sheets: {list(sheets.keys()) if sheets else []}")
            logger.info(f"[CREATE] Title: {title}")
            logger.info(f"[CREATE] Auto format: {auto_format}")
            logger.info(f"[CREATE] Excel library: {self.excel_lib}")
            
            if not self.excel_lib:
                logger.error("[CREATE] No Excel library available!")
                return {
                    "success": False,
                    "error": "Excel library not installed",
                    "output_path": output_path
                }
            
            # Validate output path
            logger.info(f"[CREATE] Validating output path...")
            logger.info(f"[CREATE] output_path == '.': {output_path == '.'}")
            logger.info(f"[CREATE] output_path is empty: {not output_path}")
            
            if not output_path or output_path == '.':
                error_msg = "Invalid or empty output path provided"
                logger.error(f"[CREATE] ERROR: {error_msg}")
                logger.error(f"[CREATE] Received output_path: '{output_path}'")
                return {
                    "success": False,
                    "error": error_msg,
                    "output_path": output_path
                }
            
            logger.info(f"[CREATE] Creating Path object...")
            output_path_obj = Path(output_path)
            logger.info(f"[CREATE] Path object created: {output_path_obj}")
            logger.info(f"[CREATE] Path parent: {output_path_obj.parent}")
            logger.info(f"[CREATE] Path name: {output_path_obj.name}")
            
            # Create parent directory if it doesn't exist (but only if parent is not current dir)
            if output_path_obj.parent != Path('.'):
                logger.info(f"[CREATE] Creating parent directory: {output_path_obj.parent}")
                output_path_obj.parent.mkdir(parents=True, exist_ok=True)
                logger.info(f"[CREATE] Parent directory created successfully")
            else:
                logger.info(f"[CREATE] Parent is current directory, skipping mkdir")
            
            logger.info(f"[CREATE] Creating Excel file at: {output_path}")
            logger.info(f"[CREATE] Using library: {self.excel_lib}")
            
            if self.excel_lib == "pandas":
                logger.info("[CREATE] Using pandas for Excel creation")
                import pandas as pd  # noqa: F401
                
                logger.info(f"[CREATE] Creating ExcelWriter...")
                with pd.ExcelWriter(str(output_path_obj), engine='openpyxl') as writer:
                    logger.info(f"[CREATE] ExcelWriter created, processing {len(sheets)} sheets")
                    for sheet_name, data in sheets.items():
                        logger.info(f"[CREATE] Processing sheet: '{sheet_name}'")
                        logger.info(f"[CREATE] Sheet data rows: {len(data) if isinstance(data, list) else 'not a list'}")
                        
                        if isinstance(data, list) and len(data) > 0:
                            # Assume first row is headers
                            headers = data[0]
                            rows = data[1:]
                            
                            logger.info(f"[CREATE] Headers: {headers}")
                            logger.info(f"[CREATE] Data rows: {len(rows)}")
                            
                            df = pd.DataFrame(rows, columns=headers)
                            df.to_excel(writer, sheet_name=str(sheet_name), index=False)
                            logger.info(f"[CREATE] Sheet '{sheet_name}' written successfully")
                        else:
                            logger.warning(f"[CREATE] Sheet '{sheet_name}' has no data or invalid format")
                
                logger.info(f"[CREATE] Excel file created successfully: {output_path}")
                logger.info(f"[CREATE] File size: {output_path_obj.stat().st_size} bytes")
                logger.info("=" * 80)
                
                return {
                    "success": True,
                    "output_path": str(output_path_obj),
                    "sheets_created": len(sheets),
                    "file_size": output_path_obj.stat().st_size,
                    "created_at": datetime.now().isoformat()
                }
            
            elif self.excel_lib == "openpyxl":
                logger.info("[CREATE] Using openpyxl for Excel creation")
                from openpyxl import Workbook  # noqa: F401
                from openpyxl.styles import Font, PatternFill, Alignment  # noqa: F401
                
                logger.info("[CREATE] Creating Workbook object...")
                workbook = Workbook()
                logger.info("[CREATE] Workbook created")
                
                # Remove default sheet if adding new ones
                if len(sheets) > 0:
                    logger.info(f"[CREATE] Removing default sheet, will create {len(sheets)} custom sheets")
                    if workbook.active:
                        workbook.remove(workbook.active)
                        logger.info("[CREATE] Default sheet removed")
                
                for sheet_name, data in sheets.items():
                    logger.info(f"[CREATE] Creating sheet: '{sheet_name}'")
                    ws = workbook.create_sheet(title=str(sheet_name)[:31])  # Excel limits to 31 chars
                    logger.info(f"[CREATE] Sheet created, writing {len(data)} rows")
                    
                    # Write data
                    for row_idx, row_data in enumerate(data, 1):
                        for col_idx, cell_value in enumerate(row_data, 1):
                            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                            
                            # Format header row if auto_format is True
                            if auto_format and row_idx == 1:
                                cell.font = Font(bold=True)
                                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.alignment = Alignment(horizontal="center")
                    
                    logger.info(f"[CREATE] Sheet '{sheet_name}' data written")
                    
                    # Auto-adjust column widths
                    if auto_format:
                        logger.debug(f"[CREATE] Auto-adjusting column widths for '{sheet_name}'")
                        for column in ws.columns:
                            max_length = 0
                            col_cell = column[0]
                            # Handle MergedCell objects by using get_column_letter
                            try:
                                # Try to get column_letter directly
                                column_letter = col_cell.column_letter  # type: ignore
                            except (AttributeError, ValueError):
                                # Fallback: use get_column_letter from openpyxl.utils
                                try:
                                    from openpyxl.utils import get_column_letter  # noqa: F401
                                    col = col_cell.column
                                    if col is not None:
                                        column_letter = get_column_letter(col)
                                    else:
                                        continue
                                except (AttributeError, ValueError):
                                    continue
                            
                            for cell in column:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = adjusted_width
                        
                        logger.info(f"[CREATE] Column widths adjusted for '{sheet_name}'")
                
                logger.info(f"[CREATE] Saving workbook to: {output_path_obj}")
                workbook.save(str(output_path_obj))
                logger.info(f"[CREATE] Workbook saved successfully")
                
                file_size = output_path_obj.stat().st_size
                logger.info(f"[CREATE] Excel file created successfully: {output_path}")
                logger.info(f"[CREATE] File size: {file_size} bytes")
                logger.info("=" * 80)
                
                return {
                    "success": True,
                    "output_path": str(output_path_obj),
                    "sheets_created": len(sheets),
                    "file_size": file_size,
                    "created_at": datetime.now().isoformat()
                }
            
            logger.error("[CREATE] No Excel library available!")
            logger.info("=" * 80)
            return {
                "success": False,
                "error": "No Excel library available",
                "output_path": str(output_path)
            }
        
        except Exception as e:
            logger.error("=" * 80)
            logger.error(f"[CREATE] EXCEPTION during Excel creation!")
            logger.error(f"[CREATE] Exception type: {type(e).__name__}")
            logger.error(f"[CREATE] Exception message: {str(e)}")
            logger.log_exception("Excel creation exception:", e)
            logger.error(f"[CREATE] Output path attempted: {output_path}")
            logger.error("=" * 80)
            return {
                "success": False,
                "error": str(e),
                "output_path": str(output_path)
            }
    
    
    def write_data(
        self,
        file_path: str,
        sheet_name: str,
        data: List[List[Any]],
        start_row: int = 1,
        start_col: int = 1,
        overwrite: bool = False
    ) -> Dict[str, Any]:
        """
        Write data to an existing Excel file.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name to write to
            data: Data to write (list of rows)
            start_row: Starting row (1-indexed)
            start_col: Starting column (1-indexed)
            overwrite: Whether to overwrite existing data
        
        Returns:
            Dictionary with write result
        """
        try:
            if not self.excel_lib:
                return {
                    "success": False,
                    "error": "Excel library not installed",
                    "file": file_path
                }
            
            file_path_obj = Path(file_path)
            
            if self.excel_lib == "openpyxl":
                from openpyxl import load_workbook  # noqa: F401
                from openpyxl import Workbook  # noqa: F401
                
                if not file_path_obj.exists():
                    logger.warning(f"File does not exist, creating new: {file_path}")
                    workbook = Workbook()
                    ws = workbook.active
                    if ws:
                        ws.title = sheet_name
                else:
                    workbook = load_workbook(str(file_path_obj))
                    if sheet_name not in workbook.sheetnames:
                        ws = workbook.create_sheet(title=sheet_name)
                    else:
                        ws = workbook[sheet_name]
                
                # Write data
                if ws:
                    for row_idx, row_data in enumerate(data, start=start_row):
                        for col_idx, cell_value in enumerate(row_data, start=start_col):
                            ws.cell(row=row_idx, column=col_idx, value=cell_value)
                
                workbook.save(str(file_path_obj))
                
                logger.info(f"Data written successfully: {file_path}")
                
                return {
                    "success": True,
                    "file": str(file_path_obj),
                    "sheet": sheet_name,
                    "rows_written": len(data),
                    "columns_written": len(data[0]) if data else 0,
                    "written_at": datetime.now().isoformat()
                }
            
            elif self.excel_lib == "pandas":
                import pandas as pd  # noqa: F401
                
                if not file_path_obj.exists():
                    logger.warning(f"File does not exist, creating new: {file_path}")
                    df = pd.DataFrame(data)
                    with pd.ExcelWriter(str(file_path_obj), engine='openpyxl') as writer:
                        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                else:
                    # Read existing file and append
                    with pd.ExcelWriter(str(file_path_obj), engine='openpyxl', mode='a') as writer:
                        df = pd.DataFrame(data)
                        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                
                logger.info(f"Data written successfully: {file_path}")
                
                return {
                    "success": True,
                    "file": str(file_path_obj),
                    "sheet": sheet_name,
                    "rows_written": len(data),
                    "columns_written": len(data[0]) if data else 0,
                    "written_at": datetime.now().isoformat()
                }
            
            return {
                "success": False,
                "error": "No valid Excel library",
                "file": str(file_path)
            }
        
        except Exception as e:
            logger.error(f"Error writing data: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "file": str(file_path)
            }
    
    
    def format_sheet(
        self,
        file_path: str,
        sheet_name: str,
        header_style: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Apply formatting to an Excel sheet.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name to format
            header_style: Header formatting options
        
        Returns:
            Dictionary with format result
        """
        try:
            if self.excel_lib != "openpyxl":
                return {
                    "success": False,
                    "error": "Formatting requires openpyxl library",
                    "file": file_path
                }
            
            from openpyxl import load_workbook  # noqa: F401
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # noqa: F401
            
            file_path_obj = Path(file_path)
            if not file_path_obj.exists():
                return {
                    "success": False,
                    "error": f"File not found: {file_path}",
                    "file": str(file_path_obj)
                }
            
            workbook = load_workbook(str(file_path_obj))
            if sheet_name not in workbook.sheetnames:
                return {
                    "success": False,
                    "error": f"Sheet not found: {sheet_name}",
                    "file": str(file_path_obj)
                }
            
            ws = workbook[sheet_name]
            
            # Apply header formatting
            if header_style is None:
                header_style = {
                    "background_color": "366092",
                    "font_color": "FFFFFF",
                    "bold": True
                }
            
            for cell in ws[1]:  # First row
                if header_style.get("background_color"):
                    cell.fill = PatternFill(
                        start_color=header_style["background_color"],
                        end_color=header_style["background_color"],
                        fill_type="solid"
                    )
                if header_style.get("font_color"):
                    cell.font = Font(
                        bold=header_style.get("bold", False),
                        color=header_style["font_color"]
                    )
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                col_cell = column[0]
                # Handle MergedCell objects by using get_column_letter
                try:
                    # Try to get column_letter directly
                    column_letter = col_cell.column_letter  # type: ignore
                except (AttributeError, ValueError):
                    # Fallback: use get_column_letter from openpyxl.utils
                    try:
                        from openpyxl.utils import get_column_letter  # noqa: F401
                        col = col_cell.column
                        if col is not None:
                            column_letter = get_column_letter(col)
                        else:
                            continue
                    except (AttributeError, ValueError):
                        continue
                
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(str(file_path_obj))
            
            logger.info(f"Sheet formatted successfully: {sheet_name}")
            
            return {
                "success": True,
                "file": str(file_path_obj),
                "sheet": sheet_name,
                "formatted_at": datetime.now().isoformat()
            }
        
        except Exception as e:
            logger.error(f"Error formatting sheet: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "file": str(file_path)
            }
    
    
    def execute_task(
        self,
        *args: Any,
        **kwargs: Any
    ) -> Union[Dict[str, Any], AgentExecutionResponse]:
        """
        Execute an Excel operation with dual-format support.
        
        Supports three calling conventions:
        1. Legacy positional: execute_task(operation, parameters)
        2. Legacy dict: execute_task({'operation': ..., 'parameters': ...})
        3. Standardized: execute_task(AgentExecutionRequest)
        
        Args:
            *args: Variable positional arguments
            **kwargs: Variable keyword arguments
        
        Returns:
            Legacy dict OR AgentExecutionResponse based on input format
        """
        start_time = time.time()
        return_legacy = True
        operation = None
        parameters = None
        task_dict = None
        
        # Detect calling convention
        # Positional arguments (operation, parameters)
        if len(args) == 2:
            operation, parameters = args
            task_dict = {"operation": operation, "parameters": parameters}
            return_legacy = True
            logger.debug("Legacy positional call")
        
        # Single dict argument
        elif len(args) == 1 and isinstance(args[0], dict):
            task_dict = args[0]
            # Check if standardized request (has task_id and task_description)
            if "task_id" in task_dict and "task_description" in task_dict:
                return_legacy = False
                logger.debug(f"Standardized request call: task_id={task_dict.get('task_id')}")
            else:
                return_legacy = True
                logger.debug("Legacy dict call")
            operation = task_dict.get("operation")
            parameters = task_dict.get("parameters", {})
        
        # Keyword arguments (operation=..., parameters=...)
        elif "operation" in kwargs:
            operation = kwargs.get("operation")
            parameters = kwargs.get("parameters", {})
            task_dict = {"operation": operation, "parameters": parameters}
            return_legacy = True
            logger.debug("Legacy keyword call")
        
        else:
            raise InvalidParameterError(
                parameter_name="task",
                message="Invalid call to execute_task. Use one of:\n"
                "  - execute_task(operation, parameters)\n"
                "  - execute_task({'operation': ..., 'parameters': ...})\n"
                "  - execute_task(AgentExecutionRequest)"
            )
        
        try:
            task_id = task_dict.get("task_id", f"excel_{int(time.time())}")  # type: ignore
            
            # Ensure parameters is not None
            if parameters is None:
                parameters = {}
            
            # Ensure operation is not None
            if operation is None:
                operation = "unknown"
            
            logger.info("=" * 80)
            logger.info(f"[EXCEL AGENT EXECUTE_TASK] Starting execution")
            logger.info("=" * 80)
            logger.info(f"[EXCEL AGENT] Operation: {operation}")
            logger.info(f"[EXCEL AGENT] Task ID: {task_id}")
            logger.info(f"[EXCEL AGENT] Return format: {'legacy' if return_legacy else 'standardized'}")
            logger.info(f"[EXCEL AGENT] Parameters received: {json.dumps(parameters, indent=2, default=str)}")
            logger.info(f"[EXCEL AGENT] Excel library available: {self.excel_lib}")
            
            # Execute the operation using existing methods
            if operation == "read":
                logger.info(f"[EXCEL AGENT] Executing READ operation")
                file_path = parameters.get('file_path', '')
                logger.info(f"[EXCEL AGENT] File path: '{file_path}'")
                logger.info(f"[EXCEL AGENT] Sheet name: {parameters.get('sheet_name')}")
                
                result = self.read_excel(
                    file_path=file_path,
                    sheet_name=parameters.get('sheet_name'),
                    header=parameters.get('header', 0),
                    index_col=parameters.get('index_col')
                )
                logger.info(f"[EXCEL AGENT] READ result: success={result.get('success')}, error={result.get('error')}")
            
            elif operation == "create":
                logger.info(f"[EXCEL AGENT] Executing CREATE operation")
                
                # Support both 'output_path' and 'file_path' parameter names for flexibility
                output_path_raw = parameters.get('output_path')
                file_path_raw = parameters.get('file_path')
                folder_path = parameters.get('folder_path')
                file_name = parameters.get('file_name')
                template_name = parameters.get('template_name')
                
                logger.info(f"[EXCEL AGENT] Raw output_path: '{output_path_raw}'")
                logger.info(f"[EXCEL AGENT] Raw file_path: '{file_path_raw}'")
                logger.info(f"[EXCEL AGENT] Folder path: '{folder_path}'")
                logger.info(f"[EXCEL AGENT] File name: '{file_name}'")
                logger.info(f"[EXCEL AGENT] Template name: '{template_name}'")
                
                # Build output path from folder_path and file_name if not directly provided
                output_path = output_path_raw or file_path_raw
                
                if not output_path and folder_path and file_name:
                    # Construct the path from folder and filename
                    output_path = str(Path(folder_path) / f"{file_name}.xlsx")
                    logger.info(f"[EXCEL AGENT] Constructed output_path from folder_path + file_name: '{output_path}'")
                elif not output_path and template_name:
                    # Use template name to create default path
                    output_path = str(Path('.') / f"{template_name}.xlsx")
                    logger.warning(f"[EXCEL AGENT] No folder specified, using current directory: '{output_path}'")
                
                logger.info(f"[EXCEL AGENT] Final output_path: '{output_path}'")
                
                sheets = parameters.get('sheets', {})
                logger.info(f"[EXCEL AGENT] Sheets provided: {list(sheets.keys()) if sheets else 'None'}")
                
                # If sheets is empty but headers are provided, create a sheet with headers
                if not sheets and parameters.get('headers'):
                    sheet_name = parameters.get('sheet_name', 'Sheet1')
                    headers = parameters.get('headers', [])
                    sheets = {sheet_name: [headers]}  # Single row with headers
                    logger.info(f"[EXCEL AGENT] Created sheet '{sheet_name}' from headers: {headers}")
                
                # If still no sheets, create default structure based on template_name
                if not sheets and template_name:
                    default_sheet_name = template_name.replace(' Template', '').strip() or 'Data'
                    default_headers = ['Item', 'Description', 'Details']
                    sheets = {default_sheet_name: [default_headers]}
                    logger.info(f"[EXCEL AGENT] Created default sheet '{default_sheet_name}' with headers: {default_headers}")
                
                logger.info(f"[EXCEL AGENT] Final sheets structure: {json.dumps({k: f'{len(v)} rows' for k, v in sheets.items()}, indent=2)}")
                logger.info(f"[EXCEL AGENT] Calling create_excel()...")
                
                # Ensure output_path is a string
                if not output_path:
                    error_msg = "No output path could be determined from parameters"
                    logger.error(f"[EXCEL AGENT] ERROR: {error_msg}")
                    result = {
                        "success": False,
                        "error": error_msg,
                        "output_path": ""
                    }
                else:
                    result = self.create_excel(
                        output_path=str(output_path),
                        sheets=sheets,
                        title=parameters.get('title'),
                        auto_format=parameters.get('auto_format', True)
                    )
                
                logger.info(f"[EXCEL AGENT] CREATE result: success={result.get('success')}, error={result.get('error')}")
                logger.info(f"[EXCEL AGENT] Output path in result: '{result.get('output_path')}'")
            
            elif operation == "write":
                logger.info(f"[EXCEL AGENT] Executing WRITE operation")
                file_path = parameters.get('file_path', '')
                sheet_name = parameters.get('sheet_name', '')
                data = parameters.get('data', [])
                
                logger.info(f"[EXCEL AGENT] File path: '{file_path}'")
                logger.info(f"[EXCEL AGENT] Sheet name: '{sheet_name}'")
                logger.info(f"[EXCEL AGENT] Data rows: {len(data)}")
                logger.info(f"[EXCEL AGENT] Start row: {parameters.get('start_row', 1)}")
                logger.info(f"[EXCEL AGENT] Start col: {parameters.get('start_col', 1)}")
                
                result = self.write_data(
                    file_path=file_path,
                    sheet_name=sheet_name,
                    data=data,
                    start_row=parameters.get('start_row', 1),
                    start_col=parameters.get('start_col', 1),
                    overwrite=parameters.get('overwrite', False)
                )
                
                logger.info(f"[EXCEL AGENT] WRITE result: success={result.get('success')}, error={result.get('error')}")
            
            elif operation == "format":
                logger.info(f"[EXCEL AGENT] Executing FORMAT operation")
                file_path = parameters.get('file_path', '')
                sheet_name = parameters.get('sheet_name', '')
                
                logger.info(f"[EXCEL AGENT] File path: '{file_path}'")
                logger.info(f"[EXCEL AGENT] Sheet name: '{sheet_name}'")
                logger.info(f"[EXCEL AGENT] Header style: {parameters.get('header_style')}")
                
                result = self.format_sheet(
                    file_path=file_path,
                    sheet_name=sheet_name,
                    header_style=parameters.get('header_style')
                )
                
                logger.info(f"[EXCEL AGENT] FORMAT result: success={result.get('success')}, error={result.get('error')}")
            
            else:
                error_msg = f"Unknown operation: {operation}"
                logger.error(f"[EXCEL AGENT] ERROR: {error_msg}")
                logger.error(f"[EXCEL AGENT] Supported operations: {self.supported_operations}")
                result = {
                    "success": False,
                    "error": error_msg,
                    "supported_operations": self.supported_operations
                }
            
            logger.info("=" * 80)
            
            # Convert legacy result to standardized response
            standard_response = self._convert_to_standard_response(
                result,
                operation,
                task_id,
                start_time
            )
            
            # Publish completion event for event-driven workflows
            self._publish_completion_event(task_id, operation, standard_response)
            
            # Return in requested format
            if return_legacy:
                # Convert back to legacy format for backward compatibility
                return self._convert_to_legacy_response(standard_response)
            else:
                return standard_response
        
        except Exception as e:
            logger.error(f"Error executing Excel task: {e}", exc_info=True)
            
            # Create standardized error response
            error = exception_to_error_response(
                e,
                source=self.agent_name,
                task_id=task_dict.get("task_id", "unknown") if task_dict else "unknown"
            )
            
            error_response: AgentExecutionResponse = {
                "status": "failure",
                "success": False,
                "result": {},
                "artifacts": [],
                "execution_time_ms": int((time.time() - start_time) * 1000),
                "timestamp": datetime.now().isoformat(),
                "agent_name": self.agent_name,
                "operation": operation or "unknown",
                "blackboard_entries": [],
                "warnings": []
            }
            # Add error field separately to handle TypedDict
            error_response["error"] = error  # type: ignore
            
            if return_legacy:
                return self._convert_to_legacy_response(error_response)
            else:
                return error_response
    
    def _convert_to_standard_response(
        self,
        legacy_result: Dict[str, Any],
        operation: str,
        task_id: str,
        start_time: float
    ) -> AgentExecutionResponse:
        """Convert legacy result dict to standardized AgentExecutionResponse."""
        success = legacy_result.get("success", False)
        
        # Extract artifacts from result
        artifacts = []
        output_file = legacy_result.get("file") or legacy_result.get("output_path")
        if output_file and success:
            file_path = Path(output_file)
            if file_path.exists():
                artifacts.append({
                    "type": "xlsx",
                    "path": str(output_file),
                    "size_bytes": file_path.stat().st_size,
                    "description": f"Excel output from {operation} operation"
                })
        
        # Create blackboard entries for sharing data
        blackboard_entries = []
        if success and "data" in legacy_result:
            blackboard_entries.append({
                "key": f"excel_data_{task_id}",
                "value": legacy_result.get("data", {}),
                "scope": "workflow",
                "ttl_seconds": 3600
            })
        
        # Build standardized response
        response: AgentExecutionResponse = {
            "status": "success" if success else "failure",
            "success": success,
            "result": {
                k: v for k, v in legacy_result.items()
                if k not in ["success", "file", "output_path"]
            },
            "artifacts": artifacts,
            "execution_time_ms": int((time.time() - start_time) * 1000),
            "timestamp": datetime.now().isoformat(),
            "agent_name": self.agent_name,
            "operation": operation,
            "blackboard_entries": blackboard_entries,
            "warnings": []
        }
        
        # Add error field if present (handle TypedDict)
        if not success and "error" in legacy_result:
            response["error"] = create_error_response(  # type: ignore
                error_code="EXCEL_001",
                error_type="execution_error",
                message=legacy_result.get("error", "Unknown error"),
                source=self.agent_name
            )
        
        return response
    
    def _convert_to_legacy_response(self, standard_response: AgentExecutionResponse) -> Dict[str, Any]:
        """Convert standardized response back to legacy format for backward compatibility."""
        legacy = {
            "success": standard_response["success"],
            "result": standard_response["result"]
        }
        
        # Add file/output_path from artifacts
        if standard_response["artifacts"]:
            legacy["file"] = standard_response["artifacts"][0]["path"]
            legacy["output_path"] = standard_response["artifacts"][0]["path"]
        
        # Add error if present (use .get() for NotRequired field)
        error = standard_response.get("error")  # type: ignore
        if error:
            legacy["error"] = error["message"]  # type: ignore
        
        # Merge result fields into top level (legacy pattern)
        if isinstance(standard_response["result"], dict):
            for key, value in standard_response["result"].items():
                if key not in legacy:
                    legacy[key] = value
        
        return legacy
    
    def _publish_completion_event(
        self,
        task_id: str,
        operation: str,
        response: AgentExecutionResponse
    ):
        """Publish task completion event for event-driven workflows."""
        try:
            event = create_system_event(
                event_type="excel_processing_completed" if operation == "read" else "csv_generated",
                event_category="task_lifecycle",
                source_agent=self.agent_name,
                payload={
                    "task_id": task_id,
                    "operation": operation,
                    "success": response["success"],
                    "artifacts": response["artifacts"],
                    "blackboard_keys": [entry["key"] for entry in response["blackboard_entries"]]
                }
            )
            self.event_bus.publish(event)
            logger.debug(f"Published completion event for task {task_id}")
        except Exception as e:
            logger.warning(f"Failed to publish completion event: {e}")
